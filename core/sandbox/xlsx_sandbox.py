"""
XLSX Sandbox
============
Generates Potomac-branded Excel workbooks (.xlsx) entirely in Python using
``openpyxl``.  No Node.js, no subprocess — runs directly in-process.

The sandbox implements the full Potomac brand palette and spreadsheet
standards:
  - Yellow (#FEC00F) column headers, title block, and accent fills
  - Dark-gray (#212121) body text, Rajdhani (headings) / Quicksand (body)
  - Zebra-striped data rows, thin borders, yellow footers
  - Optional DISCLOSURES sheet

Usage
-----
    from core.sandbox.xlsx_sandbox import XlsxSandbox
    from core.file_store import store_file

    sandbox = XlsxSandbox()
    result  = sandbox.generate(spec_dict)
    if result.success:
        entry = store_file(result.data, result.filename, "xlsx", "generate_xlsx")
        download_url = f"/files/{entry.file_id}/download"
"""

from __future__ import annotations

import io
import logging
import time
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# =============================================================================
# Potomac brand constants
# =============================================================================

_YELLOW       = "FEC00F"   # Primary yellow — headers, accent
_DARK_GRAY    = "212121"   # All body text
_YELLOW_LIGHT = "FEE896"   # Subheader / subtitle fill
_ROW_ALT      = "F5F5F5"   # Alternating row background
_WHITE        = "FFFFFF"   # Primary row background
_PINK         = "EB2F5C"   # Risk / alert  (use sparingly)
_TEAL         = "00DED1"   # Investment Strategies only (use sparingly)

# Default tab colour by sheet index
_TAB_COLORS: Dict[int, str] = {
    0: _YELLOW,
    1: _DARK_GRAY,
    2: _DARK_GRAY,
    3: _PINK,
}


# =============================================================================
# Result container
# =============================================================================

class XlsxResult:
    """Lightweight result from :meth:`XlsxSandbox.generate`."""

    __slots__ = ("success", "data", "filename", "error", "exec_time_ms")

    def __init__(
        self,
        success: bool,
        data: Optional[bytes] = None,
        filename: Optional[str] = None,
        error: Optional[str] = None,
        exec_time_ms: float = 0.0,
    ) -> None:
        self.success      = success
        self.data         = data
        self.filename     = filename
        self.error        = error
        self.exec_time_ms = exec_time_ms


# =============================================================================
# XlsxSandbox
# =============================================================================

class XlsxSandbox:
    """
    Build Potomac-branded Excel workbooks from a structured spec dict.

    No LLM code generation is required — all brand styling is applied
    automatically.

    Thread safety
    -------------
    ``generate()`` creates a fresh ``Workbook`` instance per call.
    Multiple concurrent calls are safe.
    """

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(self, spec: Dict[str, Any]) -> XlsxResult:
        """
        Generate a ``.xlsx`` workbook from *spec*.

        Parameters
        ----------
        spec : dict
            Workbook specification.  Required keys: ``title``, ``sheets``.
            See the ``generate_xlsx`` tool schema for the full definition.
        """
        start = time.time()
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            # Remove the default blank sheet openpyxl always adds
            if wb.active and wb.active.title == "Sheet":
                del wb[wb.active.title]

            doc_title    = str(spec.get("title", "POTOMAC REPORT"))
            doc_subtitle = str(spec.get("subtitle", ""))
            sheets_spec  = spec.get("sheets", [])

            if not sheets_spec:
                # Flat spec — wrap in single sheet
                sheets_spec = [{
                    "name":    "SHEET1",
                    "columns": spec.get("columns", []),
                    "rows":    spec.get("rows", []),
                }]

            for idx, sheet_spec in enumerate(sheets_spec):
                name = str(sheet_spec.get("name", f"SHEET{idx + 1}")).upper()[:31]
                ws   = wb.create_sheet(name)

                # Tab colour
                tab_color = sheet_spec.get("tab_color") or _TAB_COLORS.get(idx, _DARK_GRAY)
                ws.sheet_properties.tabColor = tab_color

                columns   = [str(c) for c in sheet_spec.get("columns", [])]
                col_count = len(columns) or 1
                rows_data = sheet_spec.get("rows", [])

                # ── Title block (rows 1-2; row 3 is spacer) ───────────
                self._write_title_block(ws, doc_title, doc_subtitle, col_count)

                # ── Column headers (row 4) ─────────────────────────────
                header_row = 4
                for ci, header in enumerate(columns, start=1):
                    cell       = ws.cell(row=header_row, column=ci)
                    cell.value = header.upper()
                    self._style_header_cell(cell)
                ws.row_dimensions[header_row].height = 20

                # ── Data rows ──────────────────────────────────────────
                num_formats = sheet_spec.get("number_formats", {})
                data_start  = 5
                for ri, row_vals in enumerate(rows_data):
                    row_num = data_start + ri
                    is_alt  = ri % 2 == 1
                    for ci, val in enumerate(row_vals, start=1):
                        cell       = ws.cell(row=row_num, column=ci)
                        cell.value = val
                        self._style_data_cell(cell, is_alt)
                        fmt_key = str(ci)
                        if fmt_key in num_formats:
                            cell.number_format = num_formats[fmt_key]

                # ── Formula overrides ──────────────────────────────────
                for formula_spec in sheet_spec.get("formulas", []):
                    addr    = formula_spec.get("cell", "")
                    formula = formula_spec.get("formula", "")
                    if addr and formula:
                        ws[addr] = formula
                        self._style_data_cell(ws[addr], alt=False)

                # ── Footer row ─────────────────────────────────────────
                if sheet_spec.get("include_footer", True):
                    footer_row  = data_start + len(rows_data) + 1
                    footer_text = sheet_spec.get("footer_text") or (
                        "Potomac"
                        "  |  For Advisor Use Only"
                    )
                    self._write_footer(ws, footer_row, col_count, footer_text)

                # ── Column widths ──────────────────────────────────────
                col_widths = sheet_spec.get("col_widths", [])
                for ci, width in enumerate(col_widths, start=1):
                    ws.column_dimensions[get_column_letter(ci)].width = width
                for ci in range(len(col_widths) + 1, col_count + 1):
                    ws.column_dimensions[get_column_letter(ci)].width = 16

                # ── Merge title / subtitle ─────────────────────────────
                if col_count > 1:
                    ws.merge_cells(
                        start_row=1, start_column=1,
                        end_row=1,   end_column=col_count,
                    )
                    if doc_subtitle:
                        ws.merge_cells(
                            start_row=2, start_column=1,
                            end_row=2,   end_column=col_count,
                        )

                # ── Freeze panes ───────────────────────────────────────
                freeze = sheet_spec.get("freeze_panes", "")
                if freeze:
                    ws.freeze_panes = freeze
                elif columns:
                    ws.freeze_panes = f"A{header_row + 1}"

                # ── Charts ─────────────────────────────────────────────
                self._add_charts(ws, sheet_spec, columns, header_row, data_start, rows_data)

                # ── Conditional formatting ─────────────────────────────
                self._add_conditional_formats(ws, sheet_spec)

                # ── Excel Table ────────────────────────────────────────
                if sheet_spec.get("as_table", False):
                    self._add_excel_table(
                        ws, sheet_spec, columns, col_count,
                        header_row, data_start, rows_data,
                        get_column_letter,
                    )

                # ── Page setup ─────────────────────────────────────────
                ws.page_setup.orientation = "landscape"
                ws.page_setup.fitToWidth  = 1
                ws.page_setup.fitToHeight = 0
                ws.print_title_rows       = f"1:{header_row}"

            # ── Disclosures sheet ──────────────────────────────────────
            if spec.get("include_disclosures", True):
                disc_text = spec.get("disclosure_text") or (
                    "IMPORTANT DISCLOSURES: Past performance is not indicative "
                    "of future results. This material is for informational purposes "
                    "only and does not constitute investment advice. "
                    "Potomac | potomac.com"
                )
                self._add_disclosures_sheet(wb, disc_text)

            # ── Serialize ─────────────────────────────────────────────
            buf = io.BytesIO()
            wb.save(buf)
            data     = buf.getvalue()
            elapsed  = round((time.time() - start) * 1000, 2)
            filename = spec.get("filename") or f"{doc_title.replace(' ', '_')}.xlsx"
            logger.info(
                "XlsxSandbox ✓  %s  (%.1f KB, %.0f ms)",
                filename, len(data) / 1024, elapsed,
            )
            return XlsxResult(True, data=data, filename=filename, exec_time_ms=elapsed)

        except Exception as exc:
            logger.error("XlsxSandbox error: %s", exc, exc_info=True)
            return XlsxResult(
                False,
                error=str(exc),
                exec_time_ms=round((time.time() - start) * 1000, 2),
            )

    # ------------------------------------------------------------------
    # Styling helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _yellow_fill() -> "PatternFill":
        from openpyxl.styles import PatternFill
        return PatternFill("solid", fgColor=_YELLOW)

    @staticmethod
    def _light_yellow_fill() -> "PatternFill":
        from openpyxl.styles import PatternFill
        return PatternFill("solid", fgColor=_YELLOW_LIGHT)

    @staticmethod
    def _alt_fill() -> "PatternFill":
        from openpyxl.styles import PatternFill
        return PatternFill("solid", fgColor=_ROW_ALT)

    @staticmethod
    def _white_fill() -> "PatternFill":
        from openpyxl.styles import PatternFill
        return PatternFill("solid", fgColor=_WHITE)

    @staticmethod
    def _thin_border() -> "Border":
        from openpyxl.styles import Border, Side
        thin = Side(style="thin", color=_DARK_GRAY)
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    @staticmethod
    def _header_font(size: int = 11) -> "Font":
        """Rajdhani Bold — Potomac headline font."""
        from openpyxl.styles import Font
        return Font(name="Rajdhani", bold=True, color=_DARK_GRAY, size=size)

    @staticmethod
    def _body_font(size: int = 10) -> "Font":
        """Quicksand — Potomac body font."""
        from openpyxl.styles import Font
        return Font(name="Quicksand", color=_DARK_GRAY, size=size)

    def _write_title_block(
        self,
        ws: Any,
        title: str,
        subtitle: str,
        col_count: int,
    ) -> None:
        """Write the Potomac title block in rows 1-2 (row 3 is spacer)."""
        from openpyxl.styles import Alignment, Font

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 6   # spacer

        cell       = ws.cell(row=1, column=1)
        cell.value = title.upper()
        cell.font  = Font(name="Rajdhani", bold=True, size=16, color=_DARK_GRAY)
        cell.fill  = self._yellow_fill()
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        if subtitle:
            cell2       = ws.cell(row=2, column=1)
            cell2.value = subtitle
            cell2.font  = Font(name="Quicksand", size=10, color=_DARK_GRAY)
            cell2.fill  = self._light_yellow_fill()
            cell2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    def _style_header_cell(self, cell: Any) -> None:
        from openpyxl.styles import Alignment
        cell.font      = self._header_font()
        cell.fill      = self._yellow_fill()
        cell.border    = self._thin_border()
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def _style_data_cell(self, cell: Any, alt: bool = False) -> None:
        from openpyxl.styles import Alignment
        cell.font      = self._body_font()
        cell.fill      = self._alt_fill() if alt else self._white_fill()
        cell.border    = self._thin_border()
        cell.alignment = Alignment(vertical="center")

    def _write_footer(
        self,
        ws: Any,
        footer_row: int,
        col_count: int,
        text: str,
    ) -> None:
        from openpyxl.styles import Alignment, Font
        ws.merge_cells(
            start_row=footer_row, start_column=1,
            end_row=footer_row,   end_column=max(col_count, 1),
        )
        cell       = ws.cell(row=footer_row, column=1)
        cell.value = text
        cell.font  = Font(name="Quicksand", size=8, italic=True, color=_DARK_GRAY)
        cell.fill  = self._light_yellow_fill()
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[footer_row].height = 14

    # ------------------------------------------------------------------
    # Chart helper
    # ------------------------------------------------------------------

    @staticmethod
    def _add_charts(
        ws: Any,
        sheet_spec: Dict[str, Any],
        columns: List[str],
        header_row: int,
        data_start: int,
        rows_data: List[Any],
    ) -> None:
        from openpyxl.chart import (
            AreaChart, BarChart, LineChart, PieChart, ScatterChart, Reference,
        )

        chart_type_map = {
            "bar_chart":     BarChart,
            "line_chart":    LineChart,
            "pie_chart":     PieChart,
            "area_chart":    AreaChart,
            "scatter_chart": ScatterChart,
        }

        for chart_spec in sheet_spec.get("charts", []):
            ChartClass = chart_type_map.get(chart_spec.get("type", "bar_chart"))
            if ChartClass is None:
                continue

            chart = ChartClass()
            chart.title   = chart_spec.get("title", "")
            chart.style   = 10
            series_labels = chart_spec.get("series_labels", [])
            x_col         = chart_spec.get("x_col", 1)
            y_cols        = chart_spec.get("y_cols", [])
            data_end      = data_start + len(rows_data) - 1

            chart.x_axis.title = columns[x_col - 1] if x_col <= len(columns) else ""
            chart.y_axis.title = series_labels[0] if series_labels else ""

            if x_col >= 1:
                chart.set_categories(
                    Reference(ws, min_col=x_col, min_row=data_start, max_row=data_end)
                )

            for si, y_col in enumerate(y_cols):
                series = Reference(
                    ws, min_col=y_col, min_row=header_row, max_row=data_end
                )
                chart.append(series)
                if si < len(series_labels):
                    chart.series[si].title = series_labels[si]

            anchor = chart_spec.get("anchor", "G5")
            ws.add_chart(chart, anchor)
            chart.width  = chart_spec.get("width", 15)
            chart.height = chart_spec.get("height", 10)

    # ------------------------------------------------------------------
    # Conditional formatting helper
    # ------------------------------------------------------------------

    @staticmethod
    def _add_conditional_formats(ws: Any, sheet_spec: Dict[str, Any]) -> None:
        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
        from openpyxl.styles import Font

        for cf in sheet_spec.get("conditional_formats", []):
            cf_type   = cf.get("type")
            range_str = cf.get("range", "")
            if not range_str:
                continue

            if cf_type == "color_scale":
                colors = cf.get("colors", ["EB2F5C", "FFEB84", "63BE7B"])
                ws.conditional_formatting.add(
                    range_str,
                    ColorScaleRule(
                        start_type="min",  start_color=colors[0],
                        mid_type="num",    mid_value=0, mid_color=colors[1],
                        end_type="max",    end_color=colors[2],
                    ),
                )

            elif cf_type == "data_bars":
                color = cf.get("color", _YELLOW)
                ws.conditional_formatting.add(
                    range_str,
                    DataBarRule(start_type="min", end_type="max", color=color),
                )

            elif cf_type == "highlight_negatives":
                ws.conditional_formatting.add(
                    range_str,
                    CellIsRule(
                        operator="lessThan", formula=["0"],
                        font=Font(color=_PINK),
                    ),
                )

            elif cf_type == "highlight_positives":
                ws.conditional_formatting.add(
                    range_str,
                    CellIsRule(
                        operator="greaterThan", formula=["0"],
                        font=Font(color="276221"),
                    ),
                )

    # ------------------------------------------------------------------
    # Excel Table helper
    # ------------------------------------------------------------------

    @staticmethod
    def _add_excel_table(
        ws: Any,
        sheet_spec: Dict[str, Any],
        columns: List[str],
        col_count: int,
        header_row: int,
        data_start: int,
        rows_data: List[Any],
        get_column_letter: Any,
    ) -> None:
        from openpyxl.worksheet.table import Table, TableStyleInfo

        table_name = sheet_spec.get("table_name", "DataTable")
        last_col   = get_column_letter(col_count)
        last_row   = data_start + len(rows_data) - 1

        table = Table(
            displayName=table_name,
            ref=f"A{header_row}:{last_col}{last_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        totals = sheet_spec.get("totals_row", {})
        if totals:
            table.showTotals = True
            for col_idx, func in totals.items():
                target_name = columns[int(col_idx) - 1]
                for tc in table.tableColumns:
                    if tc.name == target_name:
                        tc.totalsRowFunction = func

        ws.add_table(table)

    # ------------------------------------------------------------------
    # Disclosures sheet
    # ------------------------------------------------------------------

    @staticmethod
    def _add_disclosures_sheet(wb: Any, text: str) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("DISCLOSURES")
        ws.sheet_properties.tabColor = _DARK_GRAY

        ws.row_dimensions[1].height  = 24
        ws.row_dimensions[2].height  = 60
        ws.column_dimensions["A"].width = 120

        title_cell       = ws.cell(row=1, column=1)
        title_cell.value = "DISCLOSURES"
        title_cell.font  = Font(name="Rajdhani", bold=True, size=14, color=_DARK_GRAY)
        title_cell.fill  = PatternFill("solid", fgColor=_YELLOW)
        title_cell.alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )

        body_cell       = ws.cell(row=2, column=1)
        body_cell.value = text
        body_cell.font  = Font(name="Quicksand", size=9, color=_DARK_GRAY)
        body_cell.alignment = Alignment(wrap_text=True, vertical="top")
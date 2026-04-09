"""
XLSX Table Importer
===================

Converts an uploaded xlsx file_id → `highlight_table` section for DOCX.
Used by the `table_from_xlsx` section type in generate_docx.

Usage:
  from core.tools_v2.xlsx_table_importer import xlsx_to_table_section

  section = xlsx_to_table_section(file_id, sheet=0, range=None, header_row=True, auto_color=True)
  spec["sections"].append(section)
"""

import json
import logging
from typing import Dict, Any, Optional, List, Union
from pathlib import Path

logger = logging.getLogger(__name__)


def xlsx_to_table_section(
    file_id: str,
    sheet: Union[int, str] = 0,
    range_str: Optional[str] = None,
    header_row: bool = True,
    auto_color: bool = True,
    caption: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Resolve an xlsx file_id and extract a range as a DOCX highlight_table section.

    Args:
        file_id: UUID of uploaded xlsx file in file_store
        sheet: Sheet name or index (default 0)
        range_str: Excel range string e.g. "A1:F10" (optional, default entire sheet)
        header_row: First row is headers (default True)
        auto_color: Enable +/- auto coloring (default True)
        caption: Table caption text (optional)

    Returns:
        Dict ready to be inserted into DOCX sections array
    """
    try:
        from openpyxl import load_workbook
        from core.file_store import get_file

        entry = get_file(file_id)
        if not entry or not entry.data:
            logger.warning("xlsx_to_table: file_id %s not found", file_id)
            return {"type": "paragraph", "text": "[Table unavailable]"}

        # Load workbook from bytes
        import io
        wb = load_workbook(filename=io.BytesIO(entry.data), data_only=True, read_only=True)

        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]

        rows: List[List[Any]] = []
        headers: List[str] = []
        col_count = 0

        # Read data
        if range_str:
            # Specific range
            for row in ws[range_str]:
                row_values = [cell.value for cell in row]
                rows.append(row_values)
                col_count = max(col_count, len(row_values))
        else:
            # Entire sheet
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
                col_count = max(col_count, len(row))

        if not rows:
            logger.warning("xlsx_to_table: no rows found in file %s", file_id)
            return {"type": "paragraph", "text": "[Table is empty]"}

        # Extract headers
        if header_row:
            headers = [str(c) if c is not None else "" for c in rows[0]]
            rows = rows[1:]
        else:
            headers = [f"Column {i+1}" for i in range(col_count)]

        # Format all values to strings
        formatted_rows: List[List[str]] = []
        for row in rows:
            f_row = []
            for v in row:
                if v is None:
                    f_row.append("")
                elif isinstance(v, (int, float)):
                    # Format numbers to 2 decimal places
                    f_row.append(f"{v:.2f}" if isinstance(v, float) else f"{v}")
                else:
                    f_row.append(str(v))
            # Pad to full column count
            while len(f_row) < col_count:
                f_row.append("")
            formatted_rows.append(f_row)

        # Build highlight_table section
        section: Dict[str, Any] = {
            "type": "highlight_table",
            "headers": headers,
            "rows": formatted_rows,
            "auto_color_cols": list(range(1, col_count)) if auto_color else [],
        }

        if caption:
            section["caption"] = caption

        logger.info("xlsx_to_table: imported %d rows x %d cols from %s",
                    len(formatted_rows), col_count, file_id)

        return section

    except Exception as exc:
        logger.error("xlsx_to_table error: %s", exc, exc_info=True)
        return {"type": "paragraph", "text": f"[Could not load table: {str(exc)}]"}
